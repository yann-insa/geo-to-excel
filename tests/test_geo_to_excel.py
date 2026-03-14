"""Tests unitaires pour geo_to_excel.app — couverture cible 95%."""
import os
import sys
import math
import tempfile
import pytest

# Ajouter src au path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from geo_to_excel.app import (
    haversine,
    calc_length_km,
    coords_to_wkt_linestring,
    coords_to_wkt_point,
    _kml_color_to_hex,
    _parse_kml_coordinates,
    aci_to_hex,
    COLOR_NAME_MAP,
    ACI_COLORS,
    _detect_lyon_arrondissement,
    _extract_fo_from_layer,
    GESTIONNAIRES,
    EXCEL_HEADERS,
    COLUMN_WIDTHS,
    PROJECTIONS,
    COLORS,
    parse_kml_file,
    parse_dxf_file,
    detect_projection,
    detect_projection_display,
    build_xlsx,
    convert,
    _CancelledError,
)

# Chemins vers les vrais fichiers de test
TESTS_DIR = os.path.dirname(__file__)
KML_FILE = os.path.join(TESTS_DIR, "mermoz nord fibre.kml")
DXF_FILE = os.path.join(TESTS_DIR, "mermoz nord fibre.dxf")

HAS_KML = os.path.exists(KML_FILE)
HAS_DXF = os.path.exists(DXF_FILE)


# ══════════════════════════════════════════════════════════════
#  HAVERSINE
# ══════════════════════════════════════════════════════════════
class TestHaversine:
    def test_same_point_returns_zero(self):
        assert haversine(4.85, 45.75, 4.85, 45.75) == 0.0

    def test_paris_lyon_approx_392km(self):
        # Paris (2.3522, 48.8566) → Lyon (4.8357, 45.7640)
        d = haversine(2.3522, 48.8566, 4.8357, 45.7640)
        assert 390_000 < d < 395_000  # ~392 km en mètres

    def test_equator_one_degree(self):
        # 1 degré de longitude à l'équateur ≈ 111.32 km
        d = haversine(0, 0, 1, 0)
        assert 111_000 < d < 112_000

    def test_symmetry(self):
        d1 = haversine(0, 0, 1, 1)
        d2 = haversine(1, 1, 0, 0)
        assert abs(d1 - d2) < 0.01

    def test_antipodal(self):
        # Points diamétralement opposés ≈ demi-périmètre terrestre
        d = haversine(0, 0, 180, 0)
        assert 20_000_000 < d < 20_100_000


# ══════════════════════════════════════════════════════════════
#  CALC_LENGTH_KM
# ══════════════════════════════════════════════════════════════
class TestCalcLengthKm:
    def test_empty_list(self):
        assert calc_length_km([]) == 0

    def test_single_point(self):
        assert calc_length_km([(4.85, 45.75, 0)]) == 0

    def test_two_points(self):
        coords = [(2.3522, 48.8566, 0), (4.8357, 45.7640, 0)]
        length = calc_length_km(coords)
        assert 390 < length < 395  # ~392 km

    def test_closed_polygon(self):
        # Triangle autour de Lyon
        coords = [
            (4.85, 45.75, 0),
            (4.86, 45.76, 0),
            (4.87, 45.75, 0),
            (4.85, 45.75, 0),
        ]
        length = calc_length_km(coords)
        assert length > 0

    def test_identical_points(self):
        coords = [(4.85, 45.75, 0)] * 5
        assert calc_length_km(coords) == 0


# ══════════════════════════════════════════════════════════════
#  WKT
# ══════════════════════════════════════════════════════════════
class TestWkt:
    def test_linestring(self):
        coords = [(4.85, 45.75, 100), (4.86, 45.76, 200)]
        wkt = coords_to_wkt_linestring(coords)
        assert wkt == "LINESTRING Z (4.85 45.75 100,4.86 45.76 200)"

    def test_linestring_single_point(self):
        coords = [(1.0, 2.0, 3.0)]
        wkt = coords_to_wkt_linestring(coords)
        assert "LINESTRING Z" in wkt

    def test_point(self):
        wkt = coords_to_wkt_point(4.85, 45.75, 100)
        assert wkt == "POINT Z (4.85 45.75 100)"

    def test_point_default_alt(self):
        wkt = coords_to_wkt_point(4.85, 45.75)
        assert wkt == "POINT Z (4.85 45.75 0)"


# ══════════════════════════════════════════════════════════════
#  KML COLOR TO HEX
# ══════════════════════════════════════════════════════════════
class TestKmlColorToHex:
    def test_red(self):
        # KML: aabbggrr → ff0000ff = alpha=ff, blue=00, green=00, red=ff
        assert _kml_color_to_hex("ff0000ff") == "#FF0000"

    def test_blue(self):
        # ff ff 00 00 → red=00 green=00 blue=ff
        assert _kml_color_to_hex("ffff0000") == "#0000FF"

    def test_green(self):
        assert _kml_color_to_hex("ff00ff00") == "#00FF00"

    def test_white(self):
        assert _kml_color_to_hex("ffffffff") == "#FFFFFF"

    def test_uppercase_input(self):
        assert _kml_color_to_hex("FFFF00FF") == "#FF00FF"

    def test_empty_string(self):
        assert _kml_color_to_hex("") == ""

    def test_none(self):
        assert _kml_color_to_hex(None) == ""

    def test_too_short(self):
        assert _kml_color_to_hex("ff00") == ""

    def test_too_long(self):
        assert _kml_color_to_hex("ff0000ff00") == ""

    def test_invalid_hex(self):
        assert _kml_color_to_hex("gghhiijj") == ""

    def test_partial_invalid(self):
        assert _kml_color_to_hex("ff00ggff") == ""


# ══════════════════════════════════════════════════════════════
#  PARSE KML COORDINATES
# ══════════════════════════════════════════════════════════════
class TestParseKmlCoordinates:
    def test_simple(self):
        coords = _parse_kml_coordinates("4.85,45.75,100 4.86,45.76,200")
        assert len(coords) == 2
        assert coords[0] == (4.85, 45.75, 100.0)
        assert coords[1] == (4.86, 45.76, 200.0)

    def test_without_altitude(self):
        coords = _parse_kml_coordinates("4.85,45.75 4.86,45.76")
        assert len(coords) == 2
        assert coords[0] == (4.85, 45.75, 0.0)

    def test_trailing_semicolons(self):
        coords = _parse_kml_coordinates("4.85,45.75,0; 4.86,45.76,0;")
        assert len(coords) == 2

    def test_empty_string(self):
        assert _parse_kml_coordinates("") == []

    def test_whitespace_only(self):
        assert _parse_kml_coordinates("   \n  ") == []

    def test_corrupted_data(self):
        coords = _parse_kml_coordinates("abc,def 4.85,45.75,0")
        assert len(coords) == 1  # seul le 2ème est valide

    def test_single_value(self):
        # Un seul nombre n'est pas un couple lon,lat
        assert _parse_kml_coordinates("4.85") == []

    def test_multiline(self):
        text = """4.85,45.75,100
                  4.86,45.76,200
                  4.87,45.77,300"""
        coords = _parse_kml_coordinates(text)
        assert len(coords) == 3


# ══════════════════════════════════════════════════════════════
#  ACI TO HEX
# ══════════════════════════════════════════════════════════════
class TestAciToHex:
    def test_red(self):
        assert aci_to_hex(1) == "#FF0000"

    def test_yellow(self):
        assert aci_to_hex(2) == "#FFFF00"

    def test_green(self):
        assert aci_to_hex(3) == "#00FF00"

    def test_white(self):
        assert aci_to_hex(7) == "#FFFFFF"

    def test_zero(self):
        assert aci_to_hex(0) == ""

    def test_negative(self):
        assert aci_to_hex(-1) == ""

    def test_256(self):
        assert aci_to_hex(256) == ""

    def test_fallback_formula(self):
        # Index pas dans le dict ACI_COLORS mais dans la plage 1-255
        result = aci_to_hex(10)
        assert result.startswith("#")
        assert len(result) == 7

    def test_known_aci_entries(self):
        for aci, expected in ACI_COLORS.items():
            assert aci_to_hex(aci) == expected


# ══════════════════════════════════════════════════════════════
#  DETECT LYON ARRONDISSEMENT
# ══════════════════════════════════════════════════════════════
class TestDetectLyonArrondissement:
    def test_mermoz_8e(self):
        # Mermoz est dans le 8e arrondissement
        result = _detect_lyon_arrondissement(4.8720, 45.7330)
        assert "8E ARRONDISSEMENT" in result

    def test_terreaux_1er(self):
        result = _detect_lyon_arrondissement(4.832, 45.769)
        assert "1E ARRONDISSEMENT" in result

    def test_bellecour_2e(self):
        result = _detect_lyon_arrondissement(4.828, 45.756)
        assert "2E ARRONDISSEMENT" in result

    def test_part_dieu_3e(self):
        result = _detect_lyon_arrondissement(4.857, 45.758)
        assert "3E ARRONDISSEMENT" in result

    def test_outside_lyon(self):
        result = _detect_lyon_arrondissement(2.35, 48.85)  # Paris
        assert result == "Non renseigné"

    def test_invalid_coords(self):
        result = _detect_lyon_arrondissement(999, 999)
        assert result == "Non renseigné"

    def test_edge_of_zone(self):
        # Juste en dehors de la zone Lyon
        result = _detect_lyon_arrondissement(4.74, 45.69)
        assert result == "Non renseigné"

    def test_all_arrondissements_reachable(self):
        from geo_to_excel.app import _LYON_ARR_CENTERS
        for arr, (lon, lat) in _LYON_ARR_CENTERS.items():
            result = _detect_lyon_arrondissement(lon, lat)
            assert f"{arr}E ARRONDISSEMENT" in result


# ══════════════════════════════════════════════════════════════
#  EXTRACT FO FROM LAYER
# ══════════════════════════════════════════════════════════════
class TestExtractFoFromLayer:
    def test_012_fo_trace(self):
        assert _extract_fo_from_layer("_012_FO_TRACE") == "12FO"

    def test_024_fo_trace(self):
        assert _extract_fo_from_layer("_024_FO_TRACE") == "24FO"

    def test_144_fo_trace(self):
        assert _extract_fo_from_layer("_144_FO_TRACE") == "144FO"

    def test_096_fo_backbone(self):
        assert _extract_fo_from_layer("_096_FO_BACKBONE") == "96FO"

    def test_no_match(self):
        assert _extract_fo_from_layer("MonLayer") == "MonLayer"

    def test_empty(self):
        assert _extract_fo_from_layer("") == ""

    def test_fo_lowercase(self):
        assert _extract_fo_from_layer("_012_fo_trace") == "12FO"

    def test_without_underscore_prefix(self):
        assert _extract_fo_from_layer("012_FO") == "12FO"


# ══════════════════════════════════════════════════════════════
#  CONSTANTES
# ══════════════════════════════════════════════════════════════
class TestConstants:
    def test_gestionnaires_not_empty(self):
        assert len(GESTIONNAIRES) >= 1

    def test_gestionnaires_contains_criter(self):
        assert "CRITER" in GESTIONNAIRES

    def test_excel_headers_29_columns(self):
        assert len(EXCEL_HEADERS) == 29

    def test_excel_headers_first_is_id(self):
        assert EXCEL_HEADERS[0] == "ID"

    def test_projections_has_lambert93(self):
        assert "EPSG:2154" in PROJECTIONS.values()

    def test_projections_has_cc46(self):
        assert "EPSG:3946" in PROJECTIONS.values()

    def test_projections_has_custom(self):
        assert "CUSTOM" in PROJECTIONS.values()

    def test_color_name_map(self):
        assert COLOR_NAME_MAP["#FF0000"] == "ROUGE"
        assert COLOR_NAME_MAP["#0000FF"] == "BLEU"

    def test_colors_theme(self):
        assert "bg_dark" in COLORS
        assert "accent" in COLORS

    def test_column_widths_has_all_columns(self):
        assert "A" in COLUMN_WIDTHS
        assert "AC" in COLUMN_WIDTHS


# ══════════════════════════════════════════════════════════════
#  PARSE KML FILE
# ══════════════════════════════════════════════════════════════
class TestParseKmlFile:
    def test_simple_kml_linestring(self):
        """Test avec un KML minimal contenant un LineString."""
        kml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark>
            <name>TestLine</name>
            <LineString>
                <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
            </LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 1
        assert rows[0]["type"] == "LINESTRING"
        assert rows[0]["name"] == "TestLine"
        assert len(rows[0]["coords"]) == 2

    def test_kml_with_point(self):
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark>
            <name>TestPoint</name>
            <Point>
                <coordinates>4.85,45.75,100</coordinates>
            </Point>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 1
        assert rows[0]["type"] == "POINT"
        assert rows[0]["latitude"] == 45.75
        assert rows[0]["longitude"] == 4.85

    def test_kml_no_namespace(self):
        """Test KML sans namespace (xmlns="" sur Document)."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document xmlns="">
        <Placemark>
            <name>NoNS</name>
            <LineString>
                <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
            </LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 1
        assert rows[0]["name"] == "NoNS"

    def test_kml_multigeometry(self):
        """Test KML avec MultiGeometry contenant plusieurs LineStrings."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark>
            <name>Multi</name>
            <MultiGeometry>
                <LineString>
                    <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
                </LineString>
                <LineString>
                    <coordinates>4.87,45.77,0 4.88,45.78,0</coordinates>
                </LineString>
            </MultiGeometry>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 2
        assert rows[0]["name"] == "Multi_1"
        assert rows[1]["name"] == "Multi_2"

    def test_kml_with_styles(self):
        """Test extraction des couleurs via Style/StyleMap."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Style id="s1">
            <LineStyle><color>ff0000ff</color></LineStyle>
        </Style>
        <Placemark>
            <name>Colored</name>
            <styleUrl>#s1</styleUrl>
            <LineString>
                <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
            </LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert rows[0]["color"] == "#FF0000"

    def test_kml_stylemap(self):
        """Test extraction des couleurs via StyleMap."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Style id="style1">
            <LineStyle><color>ffff0000</color></LineStyle>
        </Style>
        <StyleMap id="map1">
            <Pair><key>normal</key><styleUrl>#style1</styleUrl></Pair>
            <Pair><key>highlight</key><styleUrl>#style1</styleUrl></Pair>
        </StyleMap>
        <Placemark>
            <name>Mapped</name>
            <styleUrl>#map1</styleUrl>
            <LineString>
                <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
            </LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert rows[0]["color"] == "#0000FF"

    def test_kml_empty_placemark(self):
        """Placemark sans géométrie → aucune ligne."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark><name>Empty</name></Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 0

    def test_kml_linestring_one_point(self):
        """LineString avec un seul point → ignoré (besoin de ≥ 2)."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark><name>One</name>
            <LineString><coordinates>4.85,45.75,0</coordinates></LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 0

    def test_kml_corrupted(self):
        """Fichier KML mal formé → ValueError."""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False) as f:
            f.write("<not valid xml>><>")
            f.flush()
            with pytest.raises(ValueError, match="mal formé"):
                parse_kml_file(f.name)
        os.unlink(f.name)

    def test_kml_unsupported_extension(self):
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
            f.flush()
            with pytest.raises(ValueError, match="non supporté"):
                parse_kml_file(f.name)
        os.unlink(f.name)

    def test_kml_progress_callback(self):
        """Vérifie que le callback est appelé."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark><name>P</name>
            <LineString><coordinates>4.85,45.75,0 4.86,45.76,0</coordinates></LineString>
        </Placemark>
        </Document>
        </kml>"""
        calls = []
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            parse_kml_file(f.name, progress_cb=lambda msg, pct: calls.append((msg, pct)))
        os.unlink(f.name)
        assert len(calls) >= 2  # au moins 2 appels

    def test_placemark_without_name(self):
        """Placemark sans <name> → nom par défaut Segment_N."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark>
            <LineString><coordinates>4.85,45.75,0 4.86,45.76,0</coordinates></LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert rows[0]["name"] == "Segment_1"

    @pytest.mark.skipif(not HAS_KML, reason="Fichier KML de test absent")
    def test_real_kml_mermoz(self):
        """Test avec le vrai fichier KML Mermoz."""
        rows = parse_kml_file(KML_FILE)
        assert len(rows) == 20
        # Vérifie les layers
        layers = {r["layer"] for r in rows}
        assert "_024_FO_TRACE" in layers
        assert "_012_FO_TRACE" in layers
        assert "_144_FO_TRACE" in layers
        # Vérifie les couleurs
        colors = {r["color"] for r in rows}
        assert "#FF00FF" in colors  # _024_FO
        assert "#FF0000" in colors  # _144_FO
        # Vérifie les coordonnées (zone Lyon)
        for r in rows:
            lon, lat = r["coords"][0][0], r["coords"][0][1]
            assert 4.88 < lon < 4.89
            assert 45.73 < lat < 45.74


# ══════════════════════════════════════════════════════════════
#  PARSE KMZ FILE
# ══════════════════════════════════════════════════════════════
class TestParseKmzFile:
    def test_kmz_corrupted(self):
        """KMZ corrompu (pas un zip) → ValueError."""
        with tempfile.NamedTemporaryFile(suffix=".kmz", delete=False) as f:
            f.write(b"not a zip file")
            f.flush()
            with pytest.raises(ValueError, match="corrompu"):
                parse_kml_file(f.name)
        os.unlink(f.name)

    def test_kmz_no_kml_inside(self):
        """KMZ sans fichier KML dedans → ValueError."""
        import zipfile as zf
        with tempfile.NamedTemporaryFile(suffix=".kmz", delete=False) as f:
            with zf.ZipFile(f, "w") as z:
                z.writestr("readme.txt", "no kml here")
            with pytest.raises(ValueError, match="Aucun fichier KML"):
                parse_kml_file(f.name)
        os.unlink(f.name)

    def test_kmz_valid(self):
        """KMZ valide avec un KML dedans."""
        import zipfile as zf
        kml = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark><name>Z</name>
            <LineString><coordinates>4.85,45.75,0 4.86,45.76,0</coordinates></LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kmz", delete=False) as f:
            with zf.ZipFile(f, "w") as z:
                z.writestr("doc.kml", kml)
        rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 1


# ══════════════════════════════════════════════════════════════
#  PARSE DXF FILE
# ══════════════════════════════════════════════════════════════
class TestParseDxfFile:
    @pytest.mark.skipif(not HAS_DXF, reason="Fichier DXF de test absent")
    def test_real_dxf_mermoz(self):
        """Test avec le vrai fichier DXF Mermoz en CC46."""
        rows = parse_dxf_file(DXF_FILE, "EPSG:3946")
        assert len(rows) == 20
        # Vérifie les coordonnées reprojetées (zone Lyon WGS84)
        for r in rows:
            lon, lat = r["coords"][0][0], r["coords"][0][1]
            assert 4.88 < lon < 4.90, f"lon={lon} hors zone"
            assert 45.72 < lat < 45.74, f"lat={lat} hors zone"
        # Vérifie les layers
        layers = {r["layer"] for r in rows}
        assert "_144_FO_TRACE" in layers
        assert "_012_FO_TRACE" in layers

    @pytest.mark.skipif(not HAS_DXF, reason="Fichier DXF de test absent")
    def test_dxf_progress_callback(self):
        calls = []
        parse_dxf_file(DXF_FILE, "EPSG:3946",
                       progress_cb=lambda msg, pct: calls.append((msg, pct)))
        assert len(calls) >= 3

    @pytest.mark.skipif(not HAS_DXF, reason="Fichier DXF de test absent")
    def test_dxf_wgs84_passthrough(self):
        """DXF avec EPSG:4326 → pas de reprojection (passthrough)."""
        # Les coordonnées CC46 seront interprétées comme WGS84 (résultat absurde
        # mais vérifie que le passthrough fonctionne sans crash)
        rows = parse_dxf_file(DXF_FILE, "EPSG:4326")
        assert len(rows) > 0

    def test_dxf_invalid_file(self):
        """Fichier DXF invalide → ValueError."""
        with tempfile.NamedTemporaryFile(suffix=".dxf", mode="w",
                                         delete=False) as f:
            f.write("not a dxf file")
            f.flush()
            with pytest.raises(ValueError, match="invalide"):
                parse_dxf_file(f.name, "EPSG:2154")
        os.unlink(f.name)


# ══════════════════════════════════════════════════════════════
#  DETECT PROJECTION
# ══════════════════════════════════════════════════════════════
class TestDetectProjection:
    @pytest.mark.skipif(not HAS_DXF, reason="Fichier DXF de test absent")
    def test_real_dxf_cc46(self):
        """Le vrai DXF Mermoz est en CC46."""
        epsg, label, conf = detect_projection(DXF_FILE)
        assert epsg == "EPSG:3946"
        assert "CC46" in label
        assert conf == "haute"

    def test_invalid_file(self):
        """Fichier invalide → (None, None, None)."""
        with tempfile.NamedTemporaryFile(suffix=".dxf", mode="w",
                                         delete=False) as f:
            f.write("invalid")
            f.flush()
            result = detect_projection(f.name)
        os.unlink(f.name)
        assert result == (None, None, None)

    def test_detect_projection_display_success(self):
        """Vérifie le formatage de detect_projection_display."""
        # On mock detect_projection pour ce test
        import geo_to_excel.app as app
        original = app.detect_projection
        app.detect_projection = lambda p: ("EPSG:2154", "Lambert-93", "haute")
        try:
            epsg, msg = detect_projection_display("fake.dxf")
            assert epsg == "EPSG:2154"
            assert "Lambert-93" in msg
            assert "haute" in msg
        finally:
            app.detect_projection = original

    def test_detect_projection_display_failure(self):
        import geo_to_excel.app as app
        original = app.detect_projection
        app.detect_projection = lambda p: (None, None, None)
        try:
            epsg, msg = detect_projection_display("fake.dxf")
            assert epsg is None
            assert "impossible" in msg.lower() or "manuellement" in msg.lower()
        finally:
            app.detect_projection = original


# ══════════════════════════════════════════════════════════════
#  BUILD XLSX
# ══════════════════════════════════════════════════════════════
class TestBuildXlsx:
    def _make_rows(self):
        return [
            {
                "name": "Seg1", "type": "LINESTRING", "layer": "_012_FO_TRACE",
                "coords": [(4.872, 45.733, 0), (4.873, 45.734, 0)],
                "wkt": "LINESTRING Z (4.872 45.733 0,4.873 45.734 0)",
                "length": 0.142, "color": "#FF0000",
            },
            {
                "name": "Pt1", "type": "POINT", "layer": "Points",
                "coords": [(4.875, 45.735, 100)],
                "wkt": "POINT Z (4.875 45.735 100)",
                "length": 0, "color": "",
                "latitude": 45.735, "longitude": 4.875,
            },
        ]

    def test_creates_file(self):
        rows = self._make_rows()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        count = build_xlsx(rows, out, gestionnaire="CRITER")
        assert count == 2
        assert os.path.exists(out)
        assert os.path.getsize(out) > 0
        os.unlink(out)

    def test_headers_and_merge(self):
        from openpyxl import load_workbook
        rows = self._make_rows()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        build_xlsx(rows, out, gestionnaire="CRITER")
        wb = load_workbook(out)
        ws = wb["Réseau fibre"]

        # Vérifier les en-têtes fusionnées
        assert ws["A1"].value == "OUVRAGES"
        assert ws["U1"].value == "GC"

        # Vérifier les colonnes d'en-tête
        for ci, header in enumerate(EXCEL_HEADERS, 1):
            assert ws.cell(row=2, column=ci).value == header

        wb.close()
        os.unlink(out)

    def test_data_content(self):
        from openpyxl import load_workbook
        rows = self._make_rows()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        build_xlsx(rows, out, gestionnaire="VDL-Souterrain")
        wb = load_workbook(out)
        ws = wb["Réseau fibre"]

        # Ligne 3 (première donnée)
        # openpyxl lit "" comme None — les colonnes volontairement vides
        def _empty(val):
            return val in ("", None)

        assert _empty(ws.cell(row=3, column=1).value)   # ID vide
        assert ws.cell(row=3, column=2).value == "LYON VP"
        assert "ARRONDISSEMENT" in ws.cell(row=3, column=3).value
        assert ws.cell(row=3, column=4).value == "Non renseigné"
        assert ws.cell(row=3, column=5).value == "RF0001"
        assert _empty(ws.cell(row=3, column=14).value)  # PATTERN vide
        assert _empty(ws.cell(row=3, column=15).value)  # REMARQUES vide
        assert _empty(ws.cell(row=3, column=21).value)  # ELEMENT ID vide
        assert ws.cell(row=3, column=22).value == "Fibre"
        assert ws.cell(row=3, column=23).value == "VDL-Souterrain"
        assert _empty(ws.cell(row=3, column=26).value)  # COULEUR PROPRIETAIRE vide
        assert ws.cell(row=3, column=27).value == "12FO"  # LAYER
        assert ws.cell(row=3, column=28).value == "12FO"  # COUCHE
        assert ws.cell(row=3, column=29).value == "#FF0000"
        assert ws.cell(row=3, column=12).value == "ROUGE"

        # Ligne 4 (point)
        assert ws.cell(row=4, column=9).value == 45.735  # LATITUDE
        assert ws.cell(row=4, column=10).value == 4.875  # LONGITUDE
        assert ws.cell(row=4, column=5).value == "RF0002"

        wb.close()
        os.unlink(out)

    def test_empty_rows(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        count = build_xlsx([], out)
        assert count == 0
        assert os.path.exists(out)
        os.unlink(out)

    def test_progress_callback(self):
        rows = self._make_rows()
        calls = []
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        build_xlsx(rows, out, progress_cb=lambda msg, pct: calls.append(pct))
        assert len(calls) >= 2
        os.unlink(out)


# ══════════════════════════════════════════════════════════════
#  CONVERT PIPELINE
# ══════════════════════════════════════════════════════════════
class TestConvert:
    def test_kml_to_xlsx(self):
        """Conversion KML → XLSX via le pipeline complet."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark><name>L</name>
            <LineString><coordinates>4.85,45.75,0 4.86,45.76,0</coordinates></LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            kml_path = f.name
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name
        try:
            count = convert(kml_path, out_path, gestionnaire="CRITER")
            assert count == 1
            assert os.path.exists(out_path)
        finally:
            os.unlink(kml_path)
            os.unlink(out_path)

    @pytest.mark.skipif(not HAS_DXF, reason="Fichier DXF de test absent")
    def test_dxf_to_xlsx(self):
        """Conversion DXF → XLSX via le pipeline complet."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name
        try:
            count = convert(DXF_FILE, out_path, source_epsg="EPSG:3946",
                           gestionnaire="CRITER")
            assert count == 20
        finally:
            os.unlink(out_path)

    def test_dxf_without_epsg_raises(self):
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            f.write(b"fake")
            dxf_path = f.name
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name
        try:
            with pytest.raises(ValueError, match="projection"):
                convert(dxf_path, out_path, source_epsg=None)
        finally:
            os.unlink(dxf_path)
            os.unlink(out_path)

    def test_unsupported_format(self):
        with tempfile.NamedTemporaryFile(suffix=".shp", delete=False) as f:
            f.write(b"fake")
            path = f.name
        try:
            with pytest.raises(ValueError, match="non supporté"):
                convert(path, "out.xlsx")
        finally:
            os.unlink(path)

    def test_empty_file_raises(self):
        """Fichier KML vide (aucune entité) → ValueError."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document></Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            kml_path = f.name
        try:
            with pytest.raises(ValueError, match="Aucune entité"):
                convert(kml_path, "out.xlsx")
        finally:
            os.unlink(kml_path)

    def test_progress_callback(self):
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Placemark><name>X</name>
            <LineString><coordinates>4.85,45.75,0 4.86,45.76,0</coordinates></LineString>
        </Placemark>
        </Document>
        </kml>"""
        calls = []
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            kml_path = f.name
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name
        try:
            convert(kml_path, out_path,
                    progress_cb=lambda msg, pct: calls.append(pct))
            assert 100 in calls  # le dernier appel est à 100%
        finally:
            os.unlink(kml_path)
            os.unlink(out_path)

    @pytest.mark.skipif(not HAS_KML, reason="Fichier KML de test absent")
    def test_real_kml_to_xlsx(self):
        """Pipeline complet avec le vrai KML Mermoz."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name
        try:
            count = convert(KML_FILE, out_path, gestionnaire="VDL-Souterrain")
            assert count == 20

            from openpyxl import load_workbook
            wb = load_workbook(out_path)
            ws = wb["Réseau fibre"]
            # Vérifier quelques valeurs
            assert ws.cell(row=3, column=2).value == "LYON VP"
            assert ws.cell(row=3, column=23).value == "VDL-Souterrain"
            assert ws.cell(row=3, column=27).value == "24FO"  # _024_FO_TRACE
            wb.close()
        finally:
            os.unlink(out_path)


# ══════════════════════════════════════════════════════════════
#  DETECT PROJECTION — BRANCHES AVANCÉES
# ══════════════════════════════════════════════════════════════
class TestDetectProjectionBranches:
    """Tests les différentes branches de detect_projection via DXF synthétiques."""

    def _make_dxf_with_line(self, x1, y1, x2, y2):
        """Crée un DXF minimal avec une LINE aux coordonnées données."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_line((x1, y1), (x2, y2))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            return f.name

    def test_wgs84(self):
        """Coordonnées lon/lat directes → WGS84."""
        path = self._make_dxf_with_line(4.85, 45.75, 4.86, 45.76)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:4326"
        assert conf == "haute"

    def test_lambert93(self):
        """Coordonnées Lambert-93 typiques."""
        path = self._make_dxf_with_line(845000, 6520000, 846000, 6521000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:2154"

    def test_lambert1(self):
        """Coordonnées Lambert I."""
        path = self._make_dxf_with_line(600000, 800000, 601000, 801000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:27571"

    def test_lambert2(self):
        """Coordonnées Lambert II étendu."""
        path = self._make_dxf_with_line(600000, 2000000, 601000, 2001000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:27572"

    def test_lambert3(self):
        """Coordonnées Lambert III."""
        path = self._make_dxf_with_line(600000, 3400000, 601000, 3401000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:27573"

    def test_lambert4_corse(self):
        """Coordonnées Lambert IV (Corse)."""
        path = self._make_dxf_with_line(500000, 4700000, 501000, 4701000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:27574"

    def test_cc42(self):
        """Coordonnées CC42."""
        path = self._make_dxf_with_line(1600000, 1200000, 1601000, 1201000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3942"

    def test_cc46_lyon(self):
        """Coordonnées CC46 (Lyon)."""
        path = self._make_dxf_with_line(1843000, 5175000, 1844000, 5176000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3946"

    def test_cc50(self):
        """Coordonnées CC50 (Dunkerque)."""
        path = self._make_dxf_with_line(1600000, 9200000, 1601000, 9201000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3950"

    def test_utm_31n(self):
        """Coordonnées UTM 31N — zone Y qui ne chevauche pas Lambert IV."""
        # Y=5500000 est dans la plage UTM mais hors Lambert IV (4M-5.4M)
        path = self._make_dxf_with_line(450000, 5500000, 451000, 5501000)
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:32631"

    def test_empty_dxf(self):
        """DXF vide → None."""
        import ezdxf
        doc = ezdxf.new()
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg is None

    def test_dxf_with_various_entities(self):
        """DXF avec POINT, CIRCLE, INSERT (branches de detect_projection)."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        # Coordonnées CC46
        msp.add_point((1843500, 5175500))
        msp.add_circle((1843600, 5175600), radius=10)
        msp.add_blockref("block1", (1843700, 5175700))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        epsg, label, conf = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3946"


# ══════════════════════════════════════════════════════════════
#  PARSE DXF — TYPES D'ENTITÉS AVANCÉS
# ══════════════════════════════════════════════════════════════
class TestParseDxfEntities:
    """Tests de parsing DXF pour tous les types d'entités supportés."""

    def test_line(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_line((4.85, 45.75, 100), (4.86, 45.76, 200))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "LINE"
        assert len(rows[0]["coords"]) == 2

    def test_lwpolyline(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_lwpolyline([(4.85, 45.75), (4.86, 45.76), (4.87, 45.77)])
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "LWPOLYLINE"
        assert len(rows[0]["coords"]) == 3

    def test_lwpolyline_closed(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        poly = msp.add_lwpolyline([(0, 0), (1, 0), (1, 1), (0, 1)])
        poly.close()
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        # Fermé → le dernier point = le premier
        assert rows[0]["coords"][-1] == rows[0]["coords"][0]

    def test_point(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_point((4.85, 45.75, 100))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "POINT"
        assert rows[0]["latitude"] == 45.75
        assert rows[0]["longitude"] == 4.85

    def test_circle(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_circle((4.85, 45.75), radius=0.001)
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "CIRCLE"

    def test_arc(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_arc((4.85, 45.75), radius=0.01, start_angle=0, end_angle=90)
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "ARC"

    def test_ellipse(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_ellipse((4.85, 45.75), major_axis=(0.01, 0), ratio=0.5)
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "ELLIPSE"

    def test_spline(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_spline([(4.85, 45.75), (4.86, 45.76), (4.87, 45.75),
                         (4.88, 45.76)])
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "SPLINE"

    def test_insert_block(self):
        import ezdxf
        doc = ezdxf.new()
        doc.blocks.new("test_block")
        msp = doc.modelspace()
        msp.add_blockref("test_block", (4.85, 45.75, 100))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "INSERT"
        assert "test_block" in rows[0]["name"]

    def test_reprojection_cc46_to_wgs84(self):
        """Vérifie la reprojection CC46 → WGS84 donne des coordonnées Lyon."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        # Coordonnées CC46 de Mermoz
        msp.add_line((1843500, 5175000), (1843600, 5175100))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:3946")
        os.unlink(path)
        lon, lat = rows[0]["coords"][0][0], rows[0]["coords"][0][1]
        assert 4.8 < lon < 4.9, f"lon={lon}"
        assert 45.7 < lat < 45.8, f"lat={lat}"

    def test_dxf_with_layer(self):
        """Vérifie que le layer est bien extrait."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_line((0, 0), (1, 1), dxfattribs={"layer": "_012_FO_TRACE"})
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert rows[0]["layer"] == "_012_FO_TRACE"

    def test_dxf_with_color(self):
        """Vérifie que la couleur ACI est bien convertie."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_line((0, 0), (1, 1), dxfattribs={"color": 1})  # Rouge
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert rows[0]["color"] == "#FF0000"

    def test_polyline_3d(self):
        """Test POLYLINE 3D."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_polyline3d([(4.85, 45.75, 0), (4.86, 45.76, 10),
                            (4.87, 45.77, 20)])
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 1
        assert rows[0]["type"] == "POLYLINE"


# ══════════════════════════════════════════════════════════════
#  KML AVEC NAMESPACE STANDARD (pour couvrir la branche ns trouvé)
# ══════════════════════════════════════════════════════════════
class TestKmlNamespaceVariants:
    def test_kml_with_standard_ns_and_styles(self):
        """KML avec namespace standard et styles (couvre branche ns actif)."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://www.opengis.net/kml/2.2">
        <Document>
        <Style id="red">
            <LineStyle><color>ff0000ff</color></LineStyle>
        </Style>
        <StyleMap id="redmap">
            <Pair><key>normal</key><styleUrl>#red</styleUrl></Pair>
        </StyleMap>
        <Placemark>
            <name>WithNS</name>
            <styleUrl>#redmap</styleUrl>
            <LineString>
                <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
            </LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 1
        assert rows[0]["color"] == "#FF0000"

    def test_kml_google_earth_ns(self):
        """KML avec ancien namespace Google Earth."""
        kml_content = """<?xml version="1.0"?>
        <kml xmlns="http://earth.google.com/kml/2.1">
        <Document>
        <Placemark>
            <name>GE</name>
            <LineString>
                <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
            </LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 1

    def test_kml_no_root_namespace(self):
        """KML complètement sans namespace."""
        kml_content = """<?xml version="1.0"?>
        <kml>
        <Document>
        <Placemark>
            <name>NoNS</name>
            <LineString>
                <coordinates>4.85,45.75,0 4.86,45.76,0</coordinates>
            </LineString>
        </Placemark>
        </Document>
        </kml>"""
        with tempfile.NamedTemporaryFile(suffix=".kml", mode="w",
                                         delete=False, encoding="utf-8") as f:
            f.write(kml_content)
            f.flush()
            rows = parse_kml_file(f.name)
        os.unlink(f.name)
        assert len(rows) == 1


# ══════════════════════════════════════════════════════════════
#  CANCELLED ERROR
# ══════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
#  DETECT PROJECTION — ENTITÉS SUPPLÉMENTAIRES
# ══════════════════════════════════════════════════════════════
class TestDetectProjectionEntities:
    """Couvre les branches LWPOLYLINE, POLYLINE, ARC/ELLIPSE dans detect_projection."""

    def test_detect_with_lwpolyline(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_lwpolyline([(1843500, 5175000), (1843600, 5175100)])
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        epsg, _, _ = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3946"

    def test_detect_with_polyline3d(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_polyline3d([(1843500, 5175000, 0), (1843600, 5175100, 0)])
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        epsg, _, _ = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3946"

    def test_detect_with_arc(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_arc((1843500, 5175000), radius=10, start_angle=0, end_angle=90)
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        epsg, _, _ = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3946"

    def test_detect_with_ellipse(self):
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_ellipse((1843500, 5175000), major_axis=(10, 0), ratio=0.5)
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        epsg, _, _ = detect_projection(path)
        os.unlink(path)
        assert epsg == "EPSG:3946"


# ══════════════════════════════════════════════════════════════
#  PARSE DXF — CAS LIMITES
# ══════════════════════════════════════════════════════════════
class TestParseDxfEdgeCases:
    def test_lwpolyline_one_point_skipped(self):
        """LWPOLYLINE avec un seul point → ignoré."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_lwpolyline([(4.85, 45.75)])
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        assert len(rows) == 0

    def test_mixed_entities(self):
        """DXF avec plusieurs types d'entités mélangés."""
        import ezdxf
        doc = ezdxf.new()
        doc.blocks.new("b1")
        msp = doc.modelspace()
        msp.add_line((0, 0), (1, 1))
        msp.add_lwpolyline([(0, 0), (1, 0), (1, 1)])
        msp.add_point((0.5, 0.5))
        msp.add_circle((0.5, 0.5), radius=0.1)
        msp.add_arc((0.5, 0.5), radius=0.1, start_angle=0, end_angle=180)
        msp.add_blockref("b1", (0.5, 0.5))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        types = {r["type"] for r in rows}
        assert "LINE" in types
        assert "LWPOLYLINE" in types
        assert "POINT" in types
        assert "CIRCLE" in types
        assert "ARC" in types
        assert "INSERT" in types
        assert len(rows) == 6

    def test_color_default_no_color(self):
        """Entité sans couleur explicite → couleur par défaut."""
        import ezdxf
        doc = ezdxf.new()
        msp = doc.modelspace()
        msp.add_line((0, 0), (1, 1))
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as f:
            doc.saveas(f.name)
            path = f.name
        rows = parse_dxf_file(path, "EPSG:4326")
        os.unlink(path)
        # La couleur 7 (BLANC) est la couleur par défaut DXF
        assert rows[0]["color"] in ("#FFFFFF", "")


# ══════════════════════════════════════════════════════════════
#  BUILD XLSX — CAS LIMITES
# ══════════════════════════════════════════════════════════════
class TestBuildXlsxEdgeCases:
    def test_coords_outside_lyon(self):
        """Coordonnées hors Lyon → ville = 'Non renseigné'."""
        rows = [{
            "name": "P", "type": "LINESTRING", "layer": "test",
            "coords": [(2.35, 48.85, 0), (2.36, 48.86, 0)],  # Paris
            "wkt": "LINESTRING Z (2.35 48.85 0,2.36 48.86 0)",
            "length": 1.0, "color": "",
        }]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        build_xlsx(rows, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Réseau fibre"]
        assert ws.cell(row=3, column=3).value == "Non renseigné"
        wb.close()
        os.unlink(out)

    def test_no_color_no_name(self):
        """Segment sans couleur → color_name vide."""
        rows = [{
            "name": "N", "type": "LINESTRING", "layer": "test",
            "coords": [(4.872, 45.733, 0), (4.873, 45.734, 0)],
            "wkt": "LINESTRING Z (4.872 45.733 0,4.873 45.734 0)",
            "length": 0.1, "color": "",
        }]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        build_xlsx(rows, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Réseau fibre"]
        assert ws.cell(row=3, column=12).value in ("", None)  # COULEUR vide
        wb.close()
        os.unlink(out)

    def test_unknown_color(self):
        """Couleur non reconnue → color_name vide."""
        rows = [{
            "name": "U", "type": "LINESTRING", "layer": "test",
            "coords": [(4.872, 45.733, 0), (4.873, 45.734, 0)],
            "wkt": "LINESTRING Z (4.872 45.733 0,4.873 45.734 0)",
            "length": 0.1, "color": "#123456",
        }]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        build_xlsx(rows, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Réseau fibre"]
        assert ws.cell(row=3, column=12).value in ("", None)
        assert ws.cell(row=3, column=29).value == "#123456"
        wb.close()
        os.unlink(out)

    def test_empty_coords_list(self):
        """Segment avec coords vide → ville = Non renseigné."""
        rows = [{
            "name": "E", "type": "LINESTRING", "layer": "test",
            "coords": [],
            "wkt": "LINESTRING Z ()",
            "length": 0, "color": "",
        }]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        build_xlsx(rows, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Réseau fibre"]
        assert ws.cell(row=3, column=3).value == "Non renseigné"
        wb.close()
        os.unlink(out)


# ══════════════════════════════════════════════════════════════
#  CANCELLED ERROR
# ══════════════════════════════════════════════════════════════
class TestCancelledError:
    def test_is_exception(self):
        assert issubclass(_CancelledError, Exception)

    def test_can_raise_and_catch(self):
        with pytest.raises(_CancelledError):
            raise _CancelledError()
