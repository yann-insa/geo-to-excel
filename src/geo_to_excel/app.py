#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║  GEO → Excel Converter                                      ║
║  Convertit KMZ / KML / DXF en fichier Excel                 ║
║  au format « exportations ouvrages » réseau fibre (.xlsx)    ║
║                                                              ║
║  Pour les DXF : reprojection automatique vers WGS84          ║
╚══════════════════════════════════════════════════════════════╝

Dépendances :
    pip install openpyxl ezdxf pyproj

(tkinter est inclus avec Python par défaut)
"""

import tkinter as tk
from tkinter import filedialog, ttk
import xml.etree.ElementTree as ET
import zipfile
import math
import os
import sys
import threading

# ──────────────────────────────────────────────────────────────
#  THEME
# ──────────────────────────────────────────────────────────────
COLORS = {
    "bg_dark":      "#0F1117",
    "bg_card":      "#1A1D27",
    "bg_hover":     "#242836",
    "accent":       "#6C63FF",
    "accent_hover": "#8B83FF",
    "success":      "#2ECC71",
    "error":        "#E74C3C",
    "warning":      "#F39C12",
    "text_primary": "#EAEDF3",
    "text_muted":   "#6B7280",
    "border":       "#2A2E3B",
    "dropzone_bg":  "#141722",
    "dropzone_brd": "#3A3F50",
    "combo_bg":     "#1E2230",
    "combo_fg":     "#EAEDF3",
}

FONT_FAMILY = "Segoe UI"
if sys.platform == "darwin":
    FONT_FAMILY = "SF Pro Display"
elif sys.platform.startswith("linux"):
    FONT_FAMILY = "Ubuntu"

# ──────────────────────────────────────────────────────────────
#  PROJECTIONS
# ──────────────────────────────────────────────────────────────
PROJECTIONS = {
    "Lambert-93 (EPSG:2154)":                    "EPSG:2154",
    "CC42 (EPSG:3942)":                          "EPSG:3942",
    "CC43 (EPSG:3943)":                          "EPSG:3943",
    "CC44 (EPSG:3944)":                          "EPSG:3944",
    "CC45 (EPSG:3945)":                          "EPSG:3945",
    "CC46 – Lyon (EPSG:3946)":                   "EPSG:3946",
    "CC47 (EPSG:3947)":                          "EPSG:3947",
    "CC48 (EPSG:3948)":                          "EPSG:3948",
    "CC49 (EPSG:3949)":                          "EPSG:3949",
    "CC50 (EPSG:3950)":                          "EPSG:3950",
    "Lambert II étendu (EPSG:27572)":            "EPSG:27572",
    "Lambert I (EPSG:27571)":                    "EPSG:27571",
    "Lambert III (EPSG:27573)":                  "EPSG:27573",
    "Lambert IV (EPSG:27574)":                   "EPSG:27574",
    "UTM 30N (EPSG:32630)":                     "EPSG:32630",
    "UTM 31N (EPSG:32631)":                     "EPSG:32631",
    "UTM 32N (EPSG:32632)":                     "EPSG:32632",
    "RGF93 v2 / Lambert-93 (EPSG:9793)":        "EPSG:9793",
    "WGS 84 – déjà en lat/lon (EPSG:4326)":     "EPSG:4326",
    "EPSG personnalisé…":                        "CUSTOM",
}

# ──────────────────────────────────────────────────────────────
#  DÉTECTION AUTOMATIQUE DE PROJECTION (France métropolitaine)
# ──────────────────────────────────────────────────────────────
def detect_projection(dxf_path):
    """
    Détecte automatiquement le système de projection d'un DXF
    en analysant les plages de coordonnées des entités.
    Ne fonctionne que pour la France métropolitaine.

    Retourne (epsg_string, label_humain, confiance) ou (None, None, None).
    """
    try:
        import ezdxf
    except ImportError:
        return None, None, None

    try:
        doc = ezdxf.readfile(dxf_path)
    except Exception:
        return None, None, None

    msp = doc.modelspace()
    xs, ys = [], []

    for entity in msp:
        try:
            etype = entity.dxftype()
            if etype == "LINE":
                for pt in (entity.dxf.start, entity.dxf.end):
                    xs.append(pt.x); ys.append(pt.y)
            elif etype == "LWPOLYLINE":
                for p in entity.get_points(format="xyb"):
                    xs.append(p[0]); ys.append(p[1])
            elif etype == "POLYLINE":
                for p in entity.points():
                    xs.append(p[0]); ys.append(p[1])
            elif etype == "POINT":
                xs.append(entity.dxf.location.x)
                ys.append(entity.dxf.location.y)
            elif etype == "CIRCLE":
                xs.append(entity.dxf.center.x)
                ys.append(entity.dxf.center.y)
            elif etype == "INSERT":
                xs.append(entity.dxf.insert.x)
                ys.append(entity.dxf.insert.y)
            elif etype in ("ARC", "ELLIPSE"):
                xs.append(entity.dxf.center.x)
                ys.append(entity.dxf.center.y)
        except Exception:
            continue

    if not xs or not ys:
        return None, None, None

    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)
    x_med = (x_min + x_max) / 2
    y_med = (y_min + y_max) / 2

    # ── WGS84 (déjà en lat/lon) ──
    if -10 < x_min and x_max < 15 and 40 < y_min and y_max < 52:
        return "EPSG:4326", "WGS 84 (lat/lon)", "haute"

    # ── Lambert-93 (EPSG:2154) ──
    # X: ~100k–1200k, Y: ~6000k–7200k
    if 50_000 < x_med < 1_300_000 and 5_900_000 < y_med < 7_300_000:
        return "EPSG:2154", "Lambert-93", "haute"

    # ── Lambert I (EPSG:27571) ──
    # X: ~100k–1200k, Y: ~200k–1400k
    if 50_000 < x_med < 1_300_000 and 100_000 < y_med < 1_500_000:
        return "EPSG:27571", "Lambert I", "moyenne"

    # ── Lambert II étendu (EPSG:27572) ──
    # X: ~100k–1200k, Y: ~1600k–2700k
    if 50_000 < x_med < 1_300_000 and 1_500_000 < y_med < 2_800_000:
        return "EPSG:27572", "Lambert II étendu", "haute"

    # ── Lambert III (EPSG:27573) ──
    # X: ~100k–1200k, Y: ~2900k–4000k
    if 50_000 < x_med < 1_300_000 and 2_800_000 < y_med < 4_100_000:
        return "EPSG:27573", "Lambert III", "moyenne"

    # ── Lambert IV (EPSG:27574) — Corse ──
    # X: ~-650k–650k, Y: ~4100k–5300k
    if -700_000 < x_med < 700_000 and 4_000_000 < y_med < 5_400_000:
        return "EPSG:27574", "Lambert IV (Corse)", "moyenne"

    # ── Coniques Conformes (CC42–CC50) ──
    # X: ~1000k–2300k, Y varie par zone
    # CC_n : Y_centre = (n - 42) * 1_000_000 + 1_200_000
    if 900_000 < x_med < 2_400_000:
        # Identifier la zone CC par le Y médian
        # Centre CC42=1.2M, CC43=2.2M, ..., CC50=9.2M
        zone_num = round((y_med - 1_200_000) / 1_000_000) + 42
        if 42 <= zone_num <= 50:
            epsg = f"EPSG:{3900 + zone_num}"
            city_hints = {
                42: "Perpignan/Corse", 43: "Montpellier/Marseille",
                44: "Toulouse/Nice", 45: "Bordeaux/Grenoble",
                46: "Lyon/Saint-Étienne", 47: "Paris/Nantes/Tours",
                48: "Paris/Lille/Rennes", 49: "Lille/Calais",
                50: "Dunkerque",
            }
            hint = city_hints.get(zone_num, "")
            label = f"CC{zone_num} ({hint})" if hint else f"CC{zone_num}"
            return epsg, label, "haute"

    # ── UTM zones ──
    # UTM 30N: X ~290k–1600k, Y ~4500k–5800k
    # UTM 31N: X ~-200k–1100k, Y ~4500k–5800k (X centré ~500k)
    # UTM 32N: X ~-700k–600k, Y ~4500k–5800k
    if 4_400_000 < y_med < 5_900_000:
        if 200_000 < x_med < 800_000:
            # UTM : déterminer la zone par X médian
            # Zone 31 est la plus courante en France
            # On valide en reprojetant un point et en vérifiant
            # qu'il tombe en France
            try:
                from pyproj import Transformer
                for zone, epsg in [(31, "EPSG:32631"), (30, "EPSG:32630"), (32, "EPSG:32632")]:
                    t = Transformer.from_crs(epsg, "EPSG:4326", always_xy=True)
                    lon, lat = t.transform(x_med, y_med)
                    if -6 < lon < 11 and 41 < lat < 52:
                        return epsg, f"UTM {zone}N", "moyenne"
            except ImportError:
                pass

    return None, None, None


def detect_projection_display(dxf_path):
    """Version formatée pour l'affichage GUI."""
    epsg, label, confidence = detect_projection(dxf_path)
    if epsg is None:
        return None, "Détection impossible — sélectionnez manuellement"
    conf_emoji = {"haute": "🟢", "moyenne": "🟡", "basse": "🔴"}.get(confidence, "⚪")
    return epsg, f"{conf_emoji} {label} ({epsg}) — confiance {confidence}"

# ──────────────────────────────────────────────────────────────
#  UTILITAIRES GEOMETRIQUES
# ──────────────────────────────────────────────────────────────
def haversine(lon1, lat1, lon2, lat2):
    """Distance en mètres entre deux points WGS84."""
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = (math.sin(dphi / 2) ** 2
         + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def calc_length_km(coords):
    """Longueur totale d'une polyligne en km (coords = list de (lon,lat,alt))."""
    total = 0.0
    for i in range(1, len(coords)):
        total += haversine(
            coords[i - 1][0], coords[i - 1][1],
            coords[i][0], coords[i][1],
        )
    return round(total / 1000, 3)


def coords_to_wkt_linestring(coords):
    pts = [f"{lon} {lat} {alt}" for lon, lat, alt in coords]
    return f"LINESTRING Z ({','.join(pts)})"


def coords_to_wkt_point(lon, lat, alt=0):
    return f"POINT Z ({lon} {lat} {alt})"


COLOR_NAME_MAP = {
    "#FF0000": "ROUGE",  "#0000FF": "BLEU",   "#00FF00": "VERT",
    "#FFFF00": "JAUNE",  "#FF00FF": "VIOLET",  "#00FFFF": "CYAN",
    "#000000": "NOIR",   "#FFFFFF": "BLANC",
}

# AutoCAD ACI → hex (sous-ensemble courant)
ACI_COLORS = {
    1: "#FF0000",  2: "#FFFF00",  3: "#00FF00",  4: "#00FFFF",
    5: "#0000FF",  6: "#FF00FF",  7: "#FFFFFF",  8: "#808080",
    9: "#C0C0C0", 30: "#FF7F00", 40: "#FFBF00", 50: "#BFFF00",
    60: "#7FFF00", 70: "#00FF00", 80: "#00FF7F", 90: "#00FFBF",
   100: "#00FFFF",110: "#00BFFF",120: "#007FFF",130: "#0000FF",
   140: "#7F00FF",150: "#BF00FF",160: "#FF00FF",170: "#FF00BF",
   180: "#FF007F",200: "#FF3F3F",210: "#FF7F3F",
}


def aci_to_hex(aci_color):
    if aci_color in ACI_COLORS:
        return ACI_COLORS[aci_color]
    if 1 <= aci_color <= 255:
        r = (aci_color * 97) % 256
        g = (aci_color * 53) % 256
        b = (aci_color * 29) % 256
        return f"#{r:02X}{g:02X}{b:02X}"
    return ""


# ──────────────────────────────────────────────────────────────
#  PARSEUR KML / KMZ
# ──────────────────────────────────────────────────────────────
def _kml_color_to_hex(kml_color):
    """Convertit une couleur KML (aabbggrr) en #RRGGBB."""
    if not kml_color or len(kml_color) != 8:
        return ""
    # Valider que c'est bien du hexadécimal
    try:
        int(kml_color, 16)
    except ValueError:
        return ""
    rr, gg, bb = kml_color[6:8], kml_color[4:6], kml_color[2:4]
    return f"#{rr}{gg}{bb}".upper()


def _parse_kml_coordinates(coord_text):
    coords = []
    for part in coord_text.strip().split():
        parts = part.strip().rstrip(";").split(",")
        if len(parts) >= 2:
            try:
                lon, lat = float(parts[0]), float(parts[1])
                alt = float(parts[2]) if len(parts) > 2 else 0.0
                coords.append((lon, lat, alt))
            except ValueError:
                continue
    return coords


def parse_kml_file(file_path, progress_cb=None):
    """Parse un KML ou KMZ et retourne une liste de segments."""
    ext = os.path.splitext(file_path)[1].lower()

    if progress_cb:
        progress_cb("Lecture du fichier…", 10)

    if ext == ".kmz":
        try:
            with zipfile.ZipFile(file_path, "r") as z:
                kml_files = [f for f in z.namelist() if f.lower().endswith(".kml")]
                if not kml_files:
                    raise ValueError("Aucun fichier KML trouvé dans le KMZ.")
                with z.open(kml_files[0]) as fh:
                    tree = ET.parse(fh)
        except zipfile.BadZipFile:
            raise ValueError("Le fichier KMZ est corrompu ou invalide.")
    elif ext == ".kml":
        try:
            tree = ET.parse(file_path)
        except ET.ParseError as exc:
            raise ValueError(f"Le fichier KML est mal formé : {exc}")
    else:
        raise ValueError(f"Format non supporté pour le mode KML : {ext}")

    if progress_cb:
        progress_cb("Analyse des données KML…", 30)

    root = tree.getroot()
    ns = {"kml": "http://www.opengis.net/kml/2.2"}

    # Styles
    styles = {}
    for style in root.findall(".//kml:Style", ns):
        sid = style.get("id", "")
        line_el = style.find("kml:LineStyle", ns)
        color = ""
        if line_el is not None:
            col_el = line_el.find("kml:color", ns)
            if col_el is not None and col_el.text:
                color = _kml_color_to_hex(col_el.text)
        styles[sid] = color

    style_maps = {}
    for sm in root.findall(".//kml:StyleMap", ns):
        sm_id = sm.get("id", "")
        for pair in sm.findall("kml:Pair", ns):
            key_el = pair.find("kml:key", ns)
            url_el = pair.find("kml:styleUrl", ns)
            if (key_el is not None and key_el.text == "normal"
                    and url_el is not None and url_el.text):
                ref = url_el.text.lstrip("#")
                style_maps[sm_id] = styles.get(ref, "")

    placemarks = root.findall(".//kml:Placemark", ns)
    rows = []

    for idx, pm in enumerate(placemarks):
        # Nom
        name_el = pm.find("kml:n", ns) or pm.find("kml:name", ns)
        pm_name = name_el.text if name_el is not None and name_el.text else f"Segment_{idx + 1}"

        # Couleur
        style_url_el = pm.find("kml:styleUrl", ns)
        color_hex = ""
        if style_url_el is not None and style_url_el.text:
            ref = style_url_el.text.lstrip("#")
            color_hex = style_maps.get(ref, styles.get(ref, ""))

        # LineString
        ls = pm.find(".//kml:LineString", ns)
        if ls is not None:
            coord_el = ls.find("kml:coordinates", ns)
            if coord_el is not None and coord_el.text:
                coords = _parse_kml_coordinates(coord_el.text)
                if len(coords) >= 2:
                    rows.append({
                        "name": pm_name, "type": "LINESTRING",
                        "layer": pm_name, "coords": coords,
                        "wkt": coords_to_wkt_linestring(coords),
                        "length": calc_length_km(coords),
                        "color": color_hex,
                    })

        # Point
        pt = pm.find(".//kml:Point", ns)
        if pt is not None:
            coord_el = pt.find("kml:coordinates", ns)
            if coord_el is not None and coord_el.text:
                coords = _parse_kml_coordinates(coord_el.text)
                if coords:
                    lon, lat, alt = coords[0]
                    rows.append({
                        "name": pm_name, "type": "POINT",
                        "layer": pm_name, "coords": coords,
                        "wkt": coords_to_wkt_point(lon, lat, alt),
                        "length": 0,
                        "color": color_hex,
                        "latitude": lat, "longitude": lon,
                    })

    if progress_cb:
        progress_cb(f"{len(rows)} élément(s) trouvé(s)", 60)

    return rows


# ──────────────────────────────────────────────────────────────
#  PARSEUR DXF
# ──────────────────────────────────────────────────────────────
def parse_dxf_file(file_path, source_epsg, progress_cb=None):
    """Lit un DXF, extrait les entités et reprojette vers WGS84."""
    try:
        import ezdxf
    except ImportError:
        raise ImportError(
            "Le module 'ezdxf' est requis.\n"
            "Installez-le avec : pip install ezdxf"
        )
    try:
        from pyproj import Transformer
    except ImportError:
        raise ImportError(
            "Le module 'pyproj' est requis.\n"
            "Installez-le avec : pip install pyproj"
        )

    if progress_cb:
        progress_cb("Lecture du fichier DXF…", 10)

    try:
        doc = ezdxf.readfile(file_path)
    except IOError as exc:
        raise ValueError(f"Fichier DXF invalide ou corrompu : {exc}")
    except ezdxf.DXFError as exc:
        raise ValueError(f"Fichier DXF invalide ou corrompu : {exc}")
    msp = doc.modelspace()

    need_transform = (source_epsg != "EPSG:4326")
    transformer = (
        Transformer.from_crs(source_epsg, "EPSG:4326", always_xy=True)
        if need_transform else None
    )

    def reproject(x, y, z=0.0):
        if transformer:
            lon, lat = transformer.transform(x, y)
        else:
            lon, lat = x, y
        return (lon, lat, float(z))

    if progress_cb:
        progress_cb("Extraction des entités…", 25)

    entities = list(msp)
    total = len(entities)
    rows = []
    idx = 0

    def _get_color(entity):
        try:
            aci = getattr(entity.dxf, "color", 7)
            if isinstance(aci, int) and aci > 0:
                return aci_to_hex(aci)
        except (AttributeError, TypeError):
            # Attribut couleur absent ou type inattendu → défaut vide
            return ""
        return ""

    def _get_layer(entity):
        return getattr(entity.dxf, "layer", "") or ""

    def _safe_z(vec):
        """Extrait Z d'un vecteur ezdxf de façon sûre."""
        try:
            return float(vec.z)
        except (AttributeError, TypeError):
            return 0.0

    for i, entity in enumerate(entities):
        if progress_cb and total > 0 and i % max(1, total // 20) == 0:
            pct = 25 + int(50 * i / total)
            progress_cb(f"Entité {i + 1}/{total}…", pct)

        etype = entity.dxftype()
        layer = _get_layer(entity)
        color_hex = _get_color(entity)

        try:
            if etype == "LINE":
                s, e = entity.dxf.start, entity.dxf.end
                coords = [
                    reproject(s.x, s.y, _safe_z(s)),
                    reproject(e.x, e.y, _safe_z(e)),
                ]
                idx += 1
                rows.append({
                    "name": f"LINE_{idx}", "type": "LINE", "layer": layer,
                    "coords": coords,
                    "wkt": coords_to_wkt_linestring(coords),
                    "length": calc_length_km(coords),
                    "color": color_hex,
                })

            elif etype == "LWPOLYLINE":
                raw_pts = list(entity.get_points(format="xyb"))
                if len(raw_pts) < 2:
                    continue
                coords = [reproject(p[0], p[1], 0.0) for p in raw_pts]
                if entity.closed and len(coords) > 1:
                    coords.append(coords[0])
                idx += 1
                rows.append({
                    "name": f"LWPOLYLINE_{idx}", "type": "LWPOLYLINE",
                    "layer": layer, "coords": coords,
                    "wkt": coords_to_wkt_linestring(coords),
                    "length": calc_length_km(coords),
                    "color": color_hex,
                })

            elif etype == "POLYLINE":
                pts = list(entity.points())
                if len(pts) < 2:
                    continue
                coords = [
                    reproject(p[0], p[1], p[2] if len(p) > 2 else 0.0)
                    for p in pts
                ]
                if entity.is_closed and len(coords) > 1:
                    coords.append(coords[0])
                idx += 1
                rows.append({
                    "name": f"POLYLINE_{idx}", "type": "POLYLINE",
                    "layer": layer, "coords": coords,
                    "wkt": coords_to_wkt_linestring(coords),
                    "length": calc_length_km(coords),
                    "color": color_hex,
                })

            elif etype == "SPLINE":
                pts = list(entity.flattening(0.5))
                if len(pts) < 2:
                    continue
                coords = [reproject(p.x, p.y, _safe_z(p)) for p in pts]
                idx += 1
                rows.append({
                    "name": f"SPLINE_{idx}", "type": "SPLINE",
                    "layer": layer, "coords": coords,
                    "wkt": coords_to_wkt_linestring(coords),
                    "length": calc_length_km(coords),
                    "color": color_hex,
                })

            elif etype == "POINT":
                loc = entity.dxf.location
                p = reproject(loc.x, loc.y, _safe_z(loc))
                idx += 1
                rows.append({
                    "name": f"POINT_{idx}", "type": "POINT",
                    "layer": layer, "coords": [p],
                    "wkt": coords_to_wkt_point(p[0], p[1], p[2]),
                    "length": 0, "color": color_hex,
                    "latitude": p[1], "longitude": p[0],
                })

            elif etype == "CIRCLE":
                c = entity.dxf.center
                p = reproject(c.x, c.y, _safe_z(c))
                idx += 1
                rows.append({
                    "name": f"CIRCLE_{idx}", "type": "CIRCLE",
                    "layer": layer, "coords": [p],
                    "wkt": coords_to_wkt_point(p[0], p[1], p[2]),
                    "length": 0, "color": color_hex,
                    "latitude": p[1], "longitude": p[0],
                })

            elif etype == "ARC":
                pts = list(entity.flattening(1.0))
                if len(pts) < 2:
                    continue
                coords = [reproject(p.x, p.y, 0.0) for p in pts]
                idx += 1
                rows.append({
                    "name": f"ARC_{idx}", "type": "ARC",
                    "layer": layer, "coords": coords,
                    "wkt": coords_to_wkt_linestring(coords),
                    "length": calc_length_km(coords),
                    "color": color_hex,
                })

            elif etype == "ELLIPSE":
                pts = list(entity.flattening(1.0))
                if len(pts) < 2:
                    continue
                coords = [reproject(p.x, p.y, 0.0) for p in pts]
                idx += 1
                rows.append({
                    "name": f"ELLIPSE_{idx}", "type": "ELLIPSE",
                    "layer": layer, "coords": coords,
                    "wkt": coords_to_wkt_linestring(coords),
                    "length": calc_length_km(coords),
                    "color": color_hex,
                })

            elif etype == "INSERT":
                ins = entity.dxf.insert
                p = reproject(ins.x, ins.y, _safe_z(ins))
                block_name = getattr(entity.dxf, "name", "BLOCK") or "BLOCK"
                idx += 1
                rows.append({
                    "name": f"{block_name}_{idx}", "type": "INSERT",
                    "layer": layer, "coords": [p],
                    "wkt": coords_to_wkt_point(p[0], p[1], p[2]),
                    "length": 0, "color": color_hex,
                    "latitude": p[1], "longitude": p[0],
                })

        except Exception as exc:
            # Entité malformée → signaler et continuer
            _skipped = getattr(parse_dxf_file, '_skipped', 0) + 1
            parse_dxf_file._skipped = _skipped
            continue

    skipped = getattr(parse_dxf_file, '_skipped', 0)
    parse_dxf_file._skipped = 0  # reset

    if progress_cb:
        msg = f"{len(rows)} entité(s) extraite(s)"
        if skipped > 0:
            msg += f" ({skipped} ignorée(s))"
        progress_cb(msg, 75)

    return rows


# ──────────────────────────────────────────────────────────────
#  REVERSE GEOCODING ARRONDISSEMENTS LYON
# ──────────────────────────────────────────────────────────────
# Centres approximatifs des arrondissements de Lyon (WGS84)
_LYON_ARR_CENTERS = {
    1: (4.8320, 45.7690),
    2: (4.8280, 45.7560),
    3: (4.8570, 45.7580),
    4: (4.8270, 45.7740),
    5: (4.8130, 45.7590),
    6: (4.8510, 45.7710),
    7: (4.8410, 45.7400),
    8: (4.8720, 45.7330),
    9: (4.8020, 45.7790),
}


def _detect_lyon_arrondissement(lon, lat):
    """Retourne le nom de l'arrondissement de Lyon le plus proche."""
    if not (-180 <= lon <= 180 and -90 <= lat <= 90):
        return "Non renseigné"
    # Vérifier qu'on est dans la zone de Lyon (~4.75-4.95, ~45.70-45.80)
    if not (4.75 < lon < 4.95 and 45.70 < lat < 45.80):
        return "Non renseigné"
    best_arr, best_dist = 0, float("inf")
    for arr, (clon, clat) in _LYON_ARR_CENTERS.items():
        d = math.sqrt((lon - clon) ** 2 + (lat - clat) ** 2)
        if d < best_dist:
            best_dist = d
            best_arr = arr
    return f"LYON - {best_arr}E ARRONDISSEMENT"


def _extract_fo_from_layer(layer_name):
    """Extrait la capacité FO du nom de layer DXF (ex: _012_FO_TRACE → 12FO)."""
    import re
    match = re.match(r'_?(\d+)_?FO', layer_name, re.IGNORECASE)
    if match:
        fo_num = int(match.group(1))
        return f"{fo_num}FO"
    return layer_name


# ──────────────────────────────────────────────────────────────
#  GESTIONNAIRES DISPONIBLES
# ──────────────────────────────────────────────────────────────
GESTIONNAIRES = [
    "CRITER",
    "VDL-Aérien",
    "VDL-Souterrain",
    "VDL-Divers",
    "VDL-Rocade",
    "Eclairage public",
]


# ──────────────────────────────────────────────────────────────
#  EXPORT EXCEL (commun KML & DXF)
# ──────────────────────────────────────────────────────────────
EXCEL_HEADERS = [
    "ID", "SECTEUR", "VILLE", "ZONE", "NUMERO", "PARENT",
    "PARENT CLASSE D'OUVRAGES", "MULTIVILLES", "LATITUDE", "LONGITUDE",
    "COORDS WKT", "COULEUR", "LONGUEUR", "PATTERN", "REMARQUES",
    "FICHIERS", "IMAGES", "ICONE", "CONDITION ICONE", "SPECIFIQUE",
    "ELEMENT ID", "NOM ELEMENT", "GESTIONNAIRE", "COMMENTAIRE",
    "MODE DE POSE", "COULEUR PROPRIETAIRE", "LAYER", "COUCHE", "COLOR",
]

COLUMN_WIDTHS = {
    "A": 10.5,  "B": 9.3,  "C": 29.5, "D": 30.5, "E": 27,   "F": 22.3,
    "G": 29.5,  "H": 14,   "I": 14,   "J": 14,   "K": 120,  "L": 9.3,
    "M": 10.5,  "N": 9.3,  "O": 25,   "P": 10.5, "Q": 8.2,  "R": 7,
    "S": 18.7,  "T": 12.8, "U": 13,   "V": 14,   "W": 24.7, "X": 14,
    "Y": 15.3,  "Z": 24.7, "AA": 60,  "AB": 48.3,"AC": 9.3,
}


def build_xlsx(rows, output_path, gestionnaire="", progress_cb=None):
    """Construit le fichier Excel « exportations ouvrages »."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "Le module 'openpyxl' est requis.\n"
            "Installez-le avec : pip install openpyxl"
        )

    if progress_cb:
        progress_cb("Création du fichier Excel…", 80)

    wb = Workbook()
    ws = wb.active
    ws.title = "Réseau fibre"

    hdr_fill = PatternFill("solid", fgColor="FFB3AEAD")
    hdr_font = Font(name="Calibri", bold=True, color="FF000000")
    hdr_align = Alignment(horizontal="center", vertical="center")

    # Ligne 1 : en-têtes fusionnées
    ws.merge_cells("A1:T1")
    ws["A1"] = "OUVRAGES"
    ws.merge_cells("U1:AC1")
    ws["U1"] = "GC"
    for col in range(1, 30):
        c = ws.cell(row=1, column=col)
        c.fill, c.font, c.alignment = hdr_fill, hdr_font, hdr_align

    # Ligne 2 : colonnes
    for ci, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=2, column=ci)
        c.value = h
        c.fill, c.font, c.alignment = hdr_fill, hdr_font, hdr_align

    for letter, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    if progress_cb:
        progress_cb("Écriture des données…", 88)

    for ri, seg in enumerate(rows):
        r = ri + 3
        color_hex = seg.get("color", "")
        color_name = COLOR_NAME_MAP.get(color_hex.upper(), "") if color_hex else ""

        # Déterminer la ville (arrondissement Lyon) à partir des coordonnées
        coords = seg.get("coords", [])
        if coords:
            first_lon, first_lat = coords[0][0], coords[0][1]
            ville = _detect_lyon_arrondissement(first_lon, first_lat)
        else:
            ville = "Non renseigné"

        # Extraire la capacité FO du layer
        raw_layer = seg.get("layer", "")
        fo_layer = _extract_fo_from_layer(raw_layer)

        # A=ID (vide), B=SECTEUR, C=VILLE, D=ZONE, E=NUMERO
        ws.cell(row=r, column=1).value  = ""                        # ID (vide)
        ws.cell(row=r, column=2).value  = "LYON VP"                 # SECTEUR
        ws.cell(row=r, column=3).value  = ville                     # VILLE
        ws.cell(row=r, column=4).value  = "Non renseigné"           # ZONE
        ws.cell(row=r, column=5).value  = f"RF{ri + 1:04d}"         # NUMERO
        ws.cell(row=r, column=9).value  = seg.get("latitude", "")   # LATITUDE
        ws.cell(row=r, column=10).value = seg.get("longitude", "")  # LONGITUDE
        ws.cell(row=r, column=11).value = seg["wkt"]                # COORDS WKT
        ws.cell(row=r, column=12).value = color_name                # COULEUR
        ws.cell(row=r, column=13).value = seg["length"]             # LONGUEUR
        ws.cell(row=r, column=14).value = ""                        # PATTERN (vide)
        ws.cell(row=r, column=15).value = ""                        # REMARQUES (vide)
        ws.cell(row=r, column=21).value = ""                        # ELEMENT ID (vide)
        ws.cell(row=r, column=22).value = "Fibre"                   # NOM ELEMENT
        ws.cell(row=r, column=23).value = gestionnaire              # GESTIONNAIRE
        ws.cell(row=r, column=26).value = ""                        # COULEUR PROPRIETAIRE (vide)
        ws.cell(row=r, column=27).value = fo_layer                  # LAYER (12FO, 24FO...)
        ws.cell(row=r, column=28).value = fo_layer                  # COUCHE
        ws.cell(row=r, column=29).value = color_hex                 # COLOR

    if progress_cb:
        progress_cb("Sauvegarde…", 95)

    wb.save(output_path)
    return len(rows)


# ──────────────────────────────────────────────────────────────
#  PIPELINE DE CONVERSION
# ──────────────────────────────────────────────────────────────
def convert(input_path, output_path, source_epsg=None, gestionnaire="",
            progress_cb=None):
    """
    Point d'entrée unique.
    - KML / KMZ : source_epsg ignoré (déjà WGS84).
    - DXF        : source_epsg requis.
    """
    ext = os.path.splitext(input_path)[1].lower()

    if ext in (".kml", ".kmz"):
        rows = parse_kml_file(input_path, progress_cb)
    elif ext == ".dxf":
        if not source_epsg:
            raise ValueError("Un système de projection (EPSG) est requis pour les DXF.")
        rows = parse_dxf_file(input_path, source_epsg, progress_cb)
    else:
        raise ValueError(f"Format non supporté : {ext}")

    if not rows:
        raise ValueError("Aucune entité géométrique trouvée dans le fichier.")

    if len(rows) > 100_000 and progress_cb:
        progress_cb(f"⚠ {len(rows)} entités — fichier volumineux…", 78)

    count = build_xlsx(rows, output_path, gestionnaire=gestionnaire,
                       progress_cb=progress_cb)

    if progress_cb:
        progress_cb("Terminé !", 100)

    return count


# ──────────────────────────────────────────────────────────────
#  INTERFACE GRAPHIQUE
# ──────────────────────────────────────────────────────────────

class _CancelledError(Exception):
    """Levée quand l'utilisateur annule la conversion."""


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GEO → Excel Converter")
        self.configure(bg=COLORS["bg_dark"])
        self.resizable(False, False)

        win_w, win_h = 620, 720
        x = (self.winfo_screenwidth() - win_w) // 2
        y = (self.winfo_screenheight() - win_h) // 2
        self.geometry(f"{win_w}x{win_h}+{x}+{y}")

        self.selected_file = None
        self._is_dxf = False
        self._cancel_event = threading.Event()
        self._build_styles()
        self._build_ui()

    # ── Styles ttk ──
    def _build_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Dark.TCombobox",
            fieldbackground=COLORS["combo_bg"],
            background=COLORS["combo_bg"],
            foreground=COLORS["combo_fg"],
            arrowcolor=COLORS["accent"],
            bordercolor=COLORS["border"],
            lightcolor=COLORS["border"],
            darkcolor=COLORS["border"],
            selectbackground=COLORS["accent"],
            selectforeground=COLORS["text_primary"],
        )
        style.map("Dark.TCombobox",
            fieldbackground=[("readonly", COLORS["combo_bg"])],
            background=[("readonly", COLORS["combo_bg"])],
            foreground=[("readonly", COLORS["combo_fg"])],
        )
        self.option_add("*TCombobox*Listbox.background", COLORS["combo_bg"])
        self.option_add("*TCombobox*Listbox.foreground", COLORS["combo_fg"])
        self.option_add("*TCombobox*Listbox.selectBackground", COLORS["accent"])
        self.option_add("*TCombobox*Listbox.selectForeground", COLORS["text_primary"])

    # ── Construction UI ──
    def _build_ui(self):
        PAD = 32

        # Titre
        tf = tk.Frame(self, bg=COLORS["bg_dark"])
        tf.pack(fill="x", padx=PAD, pady=(24, 0))
        tk.Label(tf, text="◆", font=(FONT_FAMILY, 18),
                 fg=COLORS["accent"], bg=COLORS["bg_dark"]).pack(side="left", padx=(0, 8))
        tk.Label(tf, text="GEO → Excel", font=(FONT_FAMILY, 22, "bold"),
                 fg=COLORS["text_primary"], bg=COLORS["bg_dark"]).pack(side="left")

        tk.Label(self,
            text="Convertisseur KMZ / KML / DXF vers le format\n"
                 "« exportations ouvrages » réseau fibre (.xlsx)",
            font=(FONT_FAMILY, 10), fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"], justify="left",
        ).pack(anchor="w", padx=PAD, pady=(4, 14))

        # Dropzone
        self.drop_frame = tk.Frame(self, bg=COLORS["dropzone_bg"],
            highlightbackground=COLORS["dropzone_brd"],
            highlightthickness=2, relief="flat")
        self.drop_frame.pack(fill="x", padx=PAD, ipady=20)

        self.file_icon = tk.Label(self.drop_frame, text="📂",
            font=(FONT_FAMILY, 28), bg=COLORS["dropzone_bg"])
        self.file_icon.pack(pady=(12, 2))

        self.file_label = tk.Label(self.drop_frame,
            text="Aucun fichier sélectionné",
            font=(FONT_FAMILY, 11), fg=COLORS["text_muted"],
            bg=COLORS["dropzone_bg"])
        self.file_label.pack()

        self.browse_btn = tk.Button(self.drop_frame, text="Parcourir…",
            font=(FONT_FAMILY, 10, "bold"),
            fg=COLORS["text_primary"], bg=COLORS["accent"],
            activebackground=COLORS["accent_hover"],
            activeforeground=COLORS["text_primary"],
            relief="flat", cursor="hand2", padx=20, pady=5,
            command=self._browse)
        self.browse_btn.pack(pady=(8, 12))

        self.info_label = tk.Label(self, text="", font=(FONT_FAMILY, 9),
            fg=COLORS["text_muted"], bg=COLORS["bg_dark"], anchor="w")
        self.info_label.pack(fill="x", padx=PAD, pady=(6, 0))

        # ── Bloc projection (masqué si KML) ──
        self.proj_frame = tk.Frame(self, bg=COLORS["bg_dark"])

        tk.Label(self.proj_frame,
            text="🌐  Système de projection du DXF :",
            font=(FONT_FAMILY, 11, "bold"), fg=COLORS["text_primary"],
            bg=COLORS["bg_dark"]).pack(anchor="w")

        # Résultat détection automatique
        self.detect_label = tk.Label(self.proj_frame, text="",
            font=(FONT_FAMILY, 10), fg=COLORS["success"],
            bg=COLORS["bg_dark"], anchor="w", wraplength=520, justify="left")
        self.detect_label.pack(fill="x", pady=(4, 6))

        tk.Label(self.proj_frame,
            text="Corrigez ci-dessous si la détection est incorrecte :",
            font=(FONT_FAMILY, 9), fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"]).pack(anchor="w", pady=(0, 4))

        self.proj_var = tk.StringVar(value="Lambert-93 (EPSG:2154)")
        self.proj_combo = ttk.Combobox(self.proj_frame,
            textvariable=self.proj_var,
            values=list(PROJECTIONS.keys()),
            state="readonly", width=48,
            font=(FONT_FAMILY, 10), style="Dark.TCombobox")
        self.proj_combo.pack(fill="x")
        self.proj_combo.bind("<<ComboboxSelected>>", self._on_proj_changed)

        # EPSG personnalisé
        self.custom_frame = tk.Frame(self.proj_frame, bg=COLORS["bg_dark"])
        tk.Label(self.custom_frame, text="Code EPSG :",
            font=(FONT_FAMILY, 10), fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"]).pack(side="left", padx=(0, 6))
        self.custom_var = tk.StringVar()
        self.custom_entry = tk.Entry(self.custom_frame,
            textvariable=self.custom_var,
            font=(FONT_FAMILY, 11), bg=COLORS["combo_bg"],
            fg=COLORS["combo_fg"], insertbackground=COLORS["text_primary"],
            relief="flat", width=12)
        self.custom_entry.pack(side="left")
        tk.Label(self.custom_frame, text="(ex: 2154, 3946…)",
            font=(FONT_FAMILY, 9), fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"]).pack(side="left", padx=(6, 0))

        # ── Bloc gestionnaire ──
        self.gest_frame = tk.Frame(self, bg=COLORS["bg_dark"])

        tk.Label(self.gest_frame,
            text="👷  Gestionnaire :",
            font=(FONT_FAMILY, 11, "bold"), fg=COLORS["text_primary"],
            bg=COLORS["bg_dark"]).pack(anchor="w")

        self.gest_var = tk.StringVar(value=GESTIONNAIRES[0])
        self.gest_combo = ttk.Combobox(self.gest_frame,
            textvariable=self.gest_var,
            values=GESTIONNAIRES,
            state="readonly", width=48,
            font=(FONT_FAMILY, 10), style="Dark.TCombobox")
        self.gest_combo.pack(fill="x", pady=(4, 0))

        # Progression
        self.prog_frame = tk.Frame(self, bg=COLORS["bg_dark"])
        self.prog_frame.pack(fill="x", padx=PAD, pady=(18, 0))
        self.prog_canvas = tk.Canvas(self.prog_frame, height=8,
            bg=COLORS["border"], highlightthickness=0)
        self.prog_canvas.pack(fill="x")
        self.prog_label = tk.Label(self.prog_frame, text="",
            font=(FONT_FAMILY, 9), fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"], anchor="w")
        self.prog_label.pack(fill="x", pady=(4, 0))

        # Boutons convertir + annuler
        btn_frame = tk.Frame(self, bg=COLORS["bg_dark"])
        btn_frame.pack(fill="x", padx=PAD, pady=(16, 0))

        self.convert_btn = tk.Button(btn_frame, text="▶  Convertir en Excel",
            font=(FONT_FAMILY, 12, "bold"),
            fg=COLORS["text_primary"], bg=COLORS["accent"],
            activebackground=COLORS["accent_hover"],
            activeforeground=COLORS["text_primary"],
            disabledforeground=COLORS["text_muted"],
            relief="flat", cursor="hand2", pady=10,
            command=self._start_convert, state="disabled")
        self.convert_btn.pack(side="left", fill="x", expand=True)

        self.cancel_btn = tk.Button(btn_frame, text="■  Annuler",
            font=(FONT_FAMILY, 12, "bold"),
            fg=COLORS["text_primary"], bg=COLORS["error"],
            activebackground="#C0392B",
            activeforeground=COLORS["text_primary"],
            disabledforeground=COLORS["text_muted"],
            relief="flat", cursor="hand2", pady=10,
            command=self._cancel_convert, state="disabled")
        self.cancel_btn.pack(side="left", fill="x", expand=True, padx=(8, 0))

        # Status
        self.status_label = tk.Label(self, text="",
            font=(FONT_FAMILY, 10), fg=COLORS["success"],
            bg=COLORS["bg_dark"], wraplength=540, justify="left")
        self.status_label.pack(fill="x", padx=PAD, pady=(10, 0))

        # Footer
        tk.Label(self,
            text="v2.0  •  KMZ / KML / DXF → Excel « exportations ouvrages »",
            font=(FONT_FAMILY, 8), fg=COLORS["border"],
            bg=COLORS["bg_dark"]).pack(side="bottom", pady=(0, 10))

    # ── Actions ──
    def _on_proj_changed(self, _event=None):
        if PROJECTIONS.get(self.proj_var.get()) == "CUSTOM":
            self.custom_frame.pack(fill="x", pady=(8, 0))
        else:
            self.custom_frame.pack_forget()

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Sélectionner un fichier KMZ, KML ou DXF",
            filetypes=[
                ("Fichiers géo", "*.kmz *.kml *.dxf"),
                ("KMZ / KML", "*.kmz *.kml"),
                ("DXF", "*.dxf"),
                ("Tous", "*.*"),
            ],
        )
        if not path:
            return

        self.selected_file = path
        ext = os.path.splitext(path)[1].lower()
        self._is_dxf = (ext == ".dxf")
        fname = os.path.basename(path)
        fsize = os.path.getsize(path)
        size_str = (f"{fsize / 1024:.1f} Ko" if fsize < 1_048_576
                    else f"{fsize / 1_048_576:.1f} Mo")

        self.file_icon.config(text="📐" if self._is_dxf else "🌍")
        self.file_label.config(text=fname, fg=COLORS["text_primary"],
                               font=(FONT_FAMILY, 11, "bold"))
        self.drop_frame.config(highlightbackground=COLORS["accent"])
        self.info_label.config(
            text=f"📍 {os.path.dirname(path)}   •   {size_str}")

        # Afficher / masquer la section projection
        if self._is_dxf:
            self.proj_frame.pack(fill="x", padx=32, pady=(14, 0),
                                 before=self.gest_frame)
            # Détection automatique de la projection
            self._run_auto_detect(path)
        else:
            self.proj_frame.pack_forget()
            self.custom_frame.pack_forget()

        # Toujours afficher le gestionnaire
        self.gest_frame.pack(fill="x", padx=32, pady=(14, 0),
                             before=self.prog_frame)

        self.convert_btn.config(state="normal")
        self.status_label.config(text="")
        self._update_progress("Prêt à convertir", 0)

    def _run_auto_detect(self, path):
        """Lance la détection de projection dans un thread séparé."""
        self.detect_label.config(
            text="⏳ Analyse des coordonnées…", fg=COLORS["text_muted"])
        self._detected_epsg = None

        def _detect():
            det_epsg, det_msg = detect_projection_display(path)
            self.after(0, lambda e=det_epsg, m=det_msg: self._on_detect_done(e, m))

        threading.Thread(target=_detect, daemon=True).start()

    def _on_detect_done(self, epsg, display_msg):
        """Callback après la détection automatique."""
        self._detected_epsg = epsg
        if epsg:
            self.detect_label.config(text=f"🔍 Détecté : {display_msg}",
                                     fg=COLORS["success"])
            # Sélectionner la bonne entrée dans le combobox
            for label, code in PROJECTIONS.items():
                if code == epsg:
                    self.proj_var.set(label)
                    self.custom_frame.pack_forget()
                    return
            # EPSG détecté mais pas dans la liste prédéfinie → custom
            self.proj_var.set("EPSG personnalisé…")
            self.custom_var.set(epsg.replace("EPSG:", ""))
            self.custom_frame.pack(fill="x", pady=(8, 0))
        else:
            self.detect_label.config(
                text="⚠️ Détection impossible — sélectionnez manuellement",
                fg=COLORS["warning"])

    def _get_epsg(self):
        selected = self.proj_var.get()
        code = PROJECTIONS.get(selected, "EPSG:2154")
        if code == "CUSTOM":
            raw = self.custom_var.get().strip()
            if not raw:
                raise ValueError("Veuillez entrer un code EPSG.")
            return raw if raw.startswith("EPSG:") else f"EPSG:{raw}"
        return code

    def _update_progress(self, text, pct):
        self.prog_label.config(text=text)
        self.prog_canvas.delete("bar")
        if pct > 0:
            w = self.prog_canvas.winfo_width()
            bw = max(1, int(w * pct / 100))
            color = COLORS["success"] if pct >= 100 else COLORS["accent"]
            self.prog_canvas.create_rectangle(
                0, 0, bw, 8, fill=color, outline="", tags="bar")

    def _start_convert(self):
        if not self.selected_file:
            return

        epsg = None
        if self._is_dxf:
            try:
                epsg = self._get_epsg()
            except ValueError as exc:
                self.status_label.config(text=f"❌  {exc}", fg=COLORS["error"])
                return
            # Valider le CRS
            try:
                from pyproj import CRS
                CRS(epsg)
            except Exception as exc:
                self.status_label.config(
                    text=f"❌  Code EPSG invalide : {epsg}\n{exc}",
                    fg=COLORS["error"])
                return

        self._cancel_event.clear()
        self.convert_btn.config(state="disabled", text="⏳  Conversion…")
        self.cancel_btn.config(state="normal")
        self.status_label.config(text="")

        thread = threading.Thread(
            target=self._do_convert, args=(epsg,), daemon=True)
        thread.start()

    def _cancel_convert(self):
        """Signale l'annulation au thread de conversion."""
        self._cancel_event.set()
        self.cancel_btn.config(state="disabled", text="■  Annulation…")
        self._update_progress("Annulation en cours…", 0)

    def _do_convert(self, epsg):
        inp = self.selected_file
        base = os.path.splitext(os.path.basename(inp))[0]
        out_dir = os.path.dirname(inp)
        gestionnaire = self.gest_var.get()

        # Vérifier les permissions d'écriture
        if not os.access(out_dir, os.W_OK):
            err_msg = f"❌  Pas de permission d'écriture dans :\n{out_dir}"
            self.after(0, lambda: self._on_done(err_msg, COLORS["error"]))
            return

        out = os.path.join(out_dir, f"{base}.xlsx")
        n = 1
        while os.path.exists(out):
            out = os.path.join(out_dir, f"{base}_{n}.xlsx")
            n += 1

        def pcb(msg, pct):
            if self._cancel_event.is_set():
                raise _CancelledError()
            self.after(0, lambda m=msg, p=pct: self._update_progress(m, p))

        try:
            count = convert(inp, out, source_epsg=epsg,
                           gestionnaire=gestionnaire, progress_cb=pcb)
            proj_info = f"\n🌐  Projection : {epsg} → WGS84" if epsg else ""
            msg_ok = (f"✅  {count} entité(s) exportée(s) avec succès !"
                      f"{proj_info}\n📁  {out}")
            self.after(0, lambda: self._on_done(msg_ok, COLORS["success"]))
        except _CancelledError:
            # Supprimer le fichier partiel
            try:
                if os.path.exists(out):
                    os.remove(out)
            except OSError:
                pass
            self.after(0, lambda: self._on_done(
                "⚠️  Conversion annulée.", COLORS["warning"]))
        except Exception as exc:
            msg_err = f"❌  Erreur : {exc}"
            self.after(0, lambda: self._on_done(msg_err, COLORS["error"]))

    def _on_done(self, msg, color):
        self.convert_btn.config(state="normal", text="▶  Convertir en Excel")
        self.cancel_btn.config(state="disabled", text="■  Annuler")
        self.status_label.config(text=msg, fg=color)
        if color == COLORS["error"]:
            self._update_progress("Erreur", 0)
        elif color == COLORS["warning"]:
            self._update_progress("Annulé", 0)


# ──────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────
def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
